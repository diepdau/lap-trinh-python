from tkinter import *
from tkinter import ttk
import re
import openpyxl
from tkinter import messagebox
from datetime import datetime

def focus1(event):
    course_field.focus_set()

def focus2(event):
    sem_field.focus_set()

def focus3(event):
    form_no_field.focus_set()

def focus4(event):
    contact_no_field.focus_set()

def focus5(event):
    email_id_field.focus_set()

def focus6(event):
    address_field.focus_set()

def register():
    # Thêm mã xử lý khi người dùng nhấp vào nút Đăng kí ở đây
    pass

def exit_program():
    root.destroy()

def clear():
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)

def is_valid_student_id(student_id):
    return re.match(r"^\d{7}$", student_id) is not None

def is_valid_email(email):
   regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b'
   return (re.fullmatch(regex, email)) is not None
def is_valid_phone_number(phone):
    return re.match(r"^\d{10}$", phone) is not None

def is_valid_semester(semester):
    return semester in ["1", "2", "3"]

def is_valid_date_of_birth(date_of_birth):
    return datetime.strptime(date_of_birth, '%d/%m/%Y') is not None

def is_valid_name(student_id):
    return re.match(r"^[a-zA-Z\s]+$", student_id) is not None

def is_valid_academic_year(academic_year):
    return re.match(r"^\d{4}-\d{4}$", academic_year) is not None

def register():
    # Validate the data
    if (not is_valid_student_id(name_field.get()) or
        not is_valid_name(course_field.get()) or
        not is_valid_date_of_birth(sem_field.get()) or
        not is_valid_email(form_no_field.get()) or
        not is_valid_phone_number(contact_no_field.get())or
        not is_valid_semester(email_id_field.get()) or
        not is_valid_academic_year(address_field.get()) or
        not any(monhoc_var[i].get() for i in range(len(mon_hoc_list)))):
        messagebox.showerror("Lỗi", "Thông tin không hợp lệ. Vui lòng kiểm tra lại.")
        return

    selected_subjects = [mon_hoc_list[i] for i in range(len(mon_hoc_list)) if monhoc_var[i].get()]

    save_to_excel(name_field.get(), course_field.get(), sem_field.get(),
                  form_no_field.get(), contact_no_field.get(), email_id_field.get(),
                  address_field.get(), selected_subjects)

    clear()
    clear_subjects()
    # Show success message
    messagebox.showinfo("Thông báo", "Đã đăng ký thành công.")


def save_to_excel(student_id, name, date_of_birth, email, phone, semester, academic_year, subjects):
    # Open or create the Excel file
    try:
        wb = openpyxl.load_workbook("excel.xlsx")
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # Select the active sheet or create a new one
    sheet = wb.active
    if sheet.title != "Đăng ký học phần":
        sheet.title = "Đăng ký học phần"

    # Find the next empty row
    row = sheet.max_row + 1

    # Write data to the Excel sheet
    sheet.cell(row=row, column=1, value=student_id)
    sheet.cell(row=row, column=2, value=name)
    sheet.cell(row=row, column=3, value=date_of_birth)
    sheet.cell(row=row, column=4, value=email)
    sheet.cell(row=row, column=5, value=phone)
    sheet.cell(row=row, column=6, value=semester)
    sheet.cell(row=row, column=7, value=academic_year)
    sheet.cell(row=row, column=8, value=", ".join(subjects))

    # Save the Excel file
    wb.save("excel.xlsx")

def clear_subjects():
    for i in range(len(mon_hoc_list)):
        monhoc_var[i].set(False)


if __name__ == "__main__":
    root = Tk()
    root.configure(background='light green')
    root.title("Form Đăng Ký Học Phần")
    root.geometry("700x300")

    heading = Label(root, text="THÔNG TIN ĐĂNG KÍ HỌC PHẦN", bg="light green",fg="red")
    name = Label(root, text="Mã số sinh viên", bg="light green")
    course = Label(root, text="Họ tên", bg="light green")
    sem = Label(root, text="Ngày sinh", bg="light green")
    form_no = Label(root, text="Email", bg="light green")
    contact_no = Label(root, text="Số điện thoại", bg="light green")
    email_id = Label(root, text="Học kỳ", bg="light green")
    address = Label(root, text="Năm học", bg="light green")
    subject = Label(root, text="Chọn môn học", bg="light green")

    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    course.grid(row=2, column=0)
    sem.grid(row=3, column=0)
    form_no.grid(row=4, column=0)
    contact_no.grid(row=5, column=0)
    email_id.grid(row=6, column=0)
    address.grid(row=7, column=0)
    subject.grid(row=8, column=0)

    nam = ["2020-2021", "2021-2022", "2022-2022", "2023-2024"]
    mon_hoc_list = ["Lập trình Python","Phát triển ứng dụng web" ,"Lập trình mạng", "TK"]
    selected_mon_hoc = []
    monhoc_var = [BooleanVar() for _ in range(len(mon_hoc_list))]
    for i, mon_hoc in enumerate(mon_hoc_list):
        checkbox = Checkbutton(root, text=mon_hoc, variable=monhoc_var[i], onvalue=True, offvalue=False, bg='light green')
        checkbox.grid(row=9+i//2, column=i%2+1, sticky=W, padx=(1,2), pady=(1,1))
    selected_nam = StringVar() #Lưu giá trị được chọn
    name_field = Entry(root)
    course_field = Entry(root)
    sem_field = Entry(root)
    form_no_field = Entry(root)
    contact_no_field = Entry(root)
    email_id_field = Entry(root)
    address_field = ttk.Combobox(root, textvariable=selected_nam, values=nam)
    

    name_field.bind("<Return>", focus1)
    course_field.bind("<Return>", focus2)
    sem_field.bind("<Return>", focus3)
    form_no_field.bind("<Return>", focus4)
    contact_no_field.bind("<Return>", focus5)
    email_id_field.bind("<Return>", focus6)
 
    
    name_field.grid(row=1, column=1, ipadx="100")
    course_field.grid(row=2, column=1, ipadx="100")
    sem_field.grid(row=3, column=1, ipadx="100")
    form_no_field.grid(row=4, column=1, ipadx="100")
    contact_no_field.grid(row=5, column=1, ipadx="100")
    email_id_field.grid(row=6, column=1, ipadx="100")
    address_field.grid(row=7, column=1, ipadx="90")
    address_field.set(nam[0])
    
    register_button = Button(root, text="Đăng kí", fg="Black",
                            bg="#3ed715",command=register)
    register_button.grid(row=12, column=0, pady=(10,0), padx=(10,5), sticky=E)

    exit_button = Button(root, text="Thoát", fg="Black",
                            bg="#3ed715",command=exit_program)
    exit_button.grid(row=12, column=1, pady=(10,0), sticky=W)

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Đăng ký học phần"

    # Ghi thông tin vào các cột
    sheet['A1'] = "Mã số sinh viên"
    sheet['B1'] = "Họ tên"
    sheet['C1'] = "Ngày sinh"
    sheet['D1'] = "Email"
    sheet['E1'] = "Số điện thoại"
    sheet['F1'] = "Học kỳ"
    sheet['G1'] = "Năm học"
    sheet['H1'] = "Môn học"

    row = sheet.max_row + 1

    # Ghi dữ liệu từ các Entry fields vào các cột tương ứng
    sheet.cell(row=row, column=1, value=name_field.get())
    sheet.cell(row=row, column=2, value=course_field.get())
    sheet.cell(row=row, column=3, value=sem_field.get())
    sheet.cell(row=row, column=4, value=form_no_field.get())
    sheet.cell(row=row, column=5, value=contact_no_field.get())
    sheet.cell(row=row, column=6, value=email_id_field.get())
    sheet.cell(row=row, column=7, value=address_field.get())

    # Lưu vào tệp excel.xlsx
    wb.save("excel.xlsx")
    root.mainloop()
